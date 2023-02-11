import time
from selenium import webdriver
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service

# Load the workbook and get the sheet
workbook = openpyxl.load_workbook('/Datadriven/web.xlsx')
sheet = workbook.active

# Start Chrome driver
browser=Service(executable_path="/usr/local/bin/chromedriver")
driver=webdriver.Chrome(service=browser)
driver.maximize_window()

url="https://demoqa.com/webtables"
driver.get(url)
time.sleep(2)

# Delete Record
driver.find_element(By.XPATH,"//span[@id='delete-record-1']").click()
driver.find_element(By.XPATH,"//span[@id='delete-record-2']").click()
driver.find_element(By.XPATH,"//span[@id='delete-record-3']").click()
time.sleep(3)

# Loop through each row in the sheet
for row in range(2, sheet.max_row + 1):
    firstname = sheet.cell(row=row, column=1).value
    lastname = sheet.cell(row=row, column=2).value
    email = sheet.cell(row=row, column=3).value
    age = sheet.cell(row=row, column=4).value
    salary = sheet.cell(row=row, column=5).value
    department = sheet.cell(row=row, column=6).value

    # scroll to location
    driver.execute_script("window.scrollBy(0,200)","")

    # click add new record button
    driver.find_element(By.XPATH,"//button[@id='addNewRecordButton']").click()
    time.sleep(3)

    # Send data
    driver.find_element(By.XPATH,"//input[@id='firstName']").send_keys(firstname)
    driver.find_element(By.XPATH,"//input[@id='lastName']").send_keys(lastname)
    driver.find_element(By.XPATH,"//input[@id='userEmail']").send_keys(email)
    driver.find_element(By.XPATH,"//input[@id='age']").send_keys(age)
    driver.find_element(By.XPATH,"//input[@id='salary']").send_keys(salary)
    driver.find_element(By.XPATH,"//input[@id='department']").send_keys(department)
    time.sleep(3)
    driver.find_element(By.XPATH,"//button[@id='submit']").click()
    time.sleep(3)

time.sleep(4)
