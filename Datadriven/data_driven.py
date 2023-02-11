import time
from selenium import webdriver
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import  Service

# Load the workbook and get the sheet
workbook = openpyxl.load_workbook('/home/vignesh/PycharmProjects/Selenium_python_demo_qa/Datadriven/datadriven.xlsx')
sheet = workbook.active

# Start Chrome driver
browser=Service(executable_path="/usr/local/bin/chromedriver")
driver=webdriver.Chrome(service=browser)
driver.maximize_window()

# Loop through each row in the sheet
for row in range(2, sheet.max_row + 1):
    user = sheet.cell(row=row, column=1).value
    pwd = sheet.cell(row=row, column=2).value


 # Navigate to the URL and enter the username and password
    driver.get('https://www.saucedemo.com/')
    time.sleep(3)
    user_1=driver.find_element(By.XPATH,"//input[@id='user-name']")
    user_1.send_keys(user)
    pass_1=driver.find_element(By.XPATH,"//input[@id='password']")
    pass_1.send_keys(pwd)
    time.sleep(4)

#Click Login
    login=driver.find_element(By.XPATH,"//input[@id='login-button']")
    login.click()

# Close the driver
driver.close()
